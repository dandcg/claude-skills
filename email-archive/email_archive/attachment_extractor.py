"""Extract text from email attachments."""

from __future__ import annotations

import io
import os

SUPPORTED_EXTENSIONS = frozenset([".pdf", ".docx", ".xlsx", ".txt", ".text", ".csv"])

EXTENSION_TO_MIME = {
    ".pdf": "application/pdf",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".doc": "application/msword",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls": "application/vnd.ms-excel",
    ".txt": "text/plain",
    ".text": "text/plain",
    ".csv": "text/csv",
}

SUPPORTED_MIMES = frozenset([
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "text/plain",
    "text/csv",
])


def get_mime_type(filename: str) -> str:
    """Get MIME type from filename extension."""
    ext = os.path.splitext(filename)[1].lower()
    return EXTENSION_TO_MIME.get(ext, "application/octet-stream")


def can_extract(filename: str, mime_type: str | None = None) -> bool:
    """Check if text can be extracted from this attachment type."""
    ext = os.path.splitext(filename)[1].lower()
    if ext in SUPPORTED_EXTENSIONS:
        return True
    if mime_type and mime_type in SUPPORTED_MIMES:
        return True
    return False


def extract_text(filename: str, content: bytes, mime_type: str | None = None) -> str | None:
    """Extract text from attachment bytes.

    Returns None if extraction fails or yields no text.
    """
    if not content:
        return None
    if not can_extract(filename, mime_type):
        return None

    ext = os.path.splitext(filename)[1].lower()
    effective_mime = mime_type or get_mime_type(filename)

    try:
        if ext == ".pdf" or effective_mime == "application/pdf":
            return _extract_pdf(content)
        elif ext == ".docx" or "wordprocessingml" in effective_mime:
            return _extract_docx(content)
        elif ext == ".xlsx" or "spreadsheetml" in effective_mime:
            return _extract_xlsx(content)
        elif ext in (".txt", ".text", ".csv") or effective_mime.startswith("text/"):
            return _extract_text(content)
        else:
            return None
    except Exception:
        return None


def _extract_pdf(content: bytes) -> str | None:
    import pdfplumber

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        pages = []
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
    text = "\n".join(pages).strip()
    return text if text else None


def _extract_docx(content: bytes) -> str | None:
    from docx import Document

    doc = Document(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs]
    text = "\n".join(paragraphs).strip()
    return text if text else None


def _extract_xlsx(content: bytes) -> str | None:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parts = []
    for ws in wb.worksheets:
        parts.append(f"--- {ws.title} ---")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            parts.append("\t".join(cells))
    wb.close()
    text = "\n".join(parts).strip()
    return text if text else None


def _extract_text(content: bytes) -> str | None:
    text = content.decode("utf-8", errors="replace").strip()
    return text if text else None
