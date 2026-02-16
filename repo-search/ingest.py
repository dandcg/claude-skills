#!/usr/bin/env python3
"""
Ingest files into ChromaDB for vector search.

Usage:
    python ingest.py [REPO_ROOT] [--db-path PATH] [--dry-run] [--force] [--verbose]

Scans for supported files (markdown, PDF, DOCX, XLSX), extracts text,
chunks them, generates embeddings, and stores in a persistent ChromaDB database.
"""

import argparse
import hashlib
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import chromadb
from langchain_text_splitters import (
    MarkdownTextSplitter,
    RecursiveCharacterTextSplitter,
)


# Supported file extensions
SUPPORTED_EXTENSIONS = {".md", ".pdf", ".docx", ".xlsx"}

# Directories to skip during scanning
SKIP_DIRS = {
    ".git", ".vectordb", "node_modules", ".venv", "__pycache__",
    "bin", "obj", ".vs", ".idea", ".claude"
}

# File patterns to skip
SKIP_FILES = {"TEMPLATE.md", "README.md"}

# Default embedding model (ChromaDB's built-in default)
DEFAULT_EMBEDDING_MODEL = "all-MiniLM-L6-v2"

# Default chunk settings
DEFAULT_CHUNK_SIZE = 1000  # characters (~250 tokens)
DEFAULT_CHUNK_OVERLAP = 200  # characters overlap between chunks

# Per-format chunk size defaults
FORMAT_CHUNK_DEFAULTS = {
    ".md": {"chunk_size": 1500, "chunk_overlap": 200},
    ".pdf": {"chunk_size": 1000, "chunk_overlap": 200},
    ".docx": {"chunk_size": 1500, "chunk_overlap": 200},
    ".xlsx": {"chunk_size": 2000, "chunk_overlap": 200},
}


def find_files(repo_root: Path) -> list[Path]:
    """Find all supported files in the repo, skipping excluded directories."""
    files = []
    for root, dirs, filenames in os.walk(repo_root):
        # Filter out skip directories in-place
        dirs[:] = [d for d in dirs if d not in SKIP_DIRS]

        for f in filenames:
            p = Path(root) / f
            if p.suffix.lower() in SUPPORTED_EXTENSIONS and f not in SKIP_FILES:
                files.append(p)

    return sorted(files)


def extract_text(file_path: Path) -> str:
    """Extract plain text from a file based on its extension."""
    ext = file_path.suffix.lower()

    if ext == ".md":
        return file_path.read_text(encoding="utf-8")
    elif ext == ".pdf":
        return _extract_pdf(file_path)
    elif ext == ".docx":
        return _extract_docx(file_path)
    elif ext == ".xlsx":
        return _extract_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def _extract_pdf(file_path: Path) -> str:
    """Extract text from a PDF file."""
    from pypdf import PdfReader

    reader = PdfReader(file_path)
    pages = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            pages.append(text)
    return "\n\n".join(pages)


def _extract_docx(file_path: Path) -> str:
    """Extract text from a Word DOCX file."""
    from docx import Document

    doc = Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def _extract_xlsx(file_path: Path) -> str:
    """Extract text from an Excel XLSX file."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        parts.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c for c in cells):
                parts.append(" | ".join(cells))
        parts.append("")
    wb.close()
    return "\n".join(parts)


def extract_metadata(file_path: Path, repo_root: Path, content: str) -> dict:
    """Extract metadata from a file's content and path."""
    rel_path = str(file_path.relative_to(repo_root))
    ext = file_path.suffix.lower()

    # Determine area from path
    parts = rel_path.split("/")
    area = parts[0] if parts else "unknown"
    sub_area = parts[1] if len(parts) > 2 else ""

    title = file_path.stem
    date = ""
    status = ""

    # Extract markdown-specific metadata
    if ext == ".md":
        title_match = re.search(r"^#\s+(.+)$", content, re.MULTILINE)
        if title_match:
            title = title_match.group(1).strip()

        date_match = re.search(
            r"\*\*(?:Added|Date|Started):\*\*\s*(\d{4}-\d{2}-\d{2})",
            content
        )
        if date_match:
            date = date_match.group(1)

        status_match = re.search(r"\*\*Status:\*\*\s*(\w+)", content)
        if status_match:
            status = status_match.group(1)

    # Try filename date pattern for all file types
    if not date:
        fname_match = re.match(r"(\d{4}-\d{2}-\d{2})", file_path.name)
        if fname_match:
            date = fname_match.group(1)

    return {
        "file_path": rel_path,
        "file_type": ext.lstrip("."),
        "area": area,
        "sub_area": sub_area,
        "title": title,
        "date": date,
        "status": status,
        "file_size": file_path.stat().st_size,
    }


def _get_heading_chain(content: str, position: int) -> str:
    """Extract the heading chain (h1 > h2 > h3) that applies at a given position."""
    lines = content[:position].split("\n")
    headings = {}
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("### "):
            headings[3] = stripped[4:].strip()
        elif stripped.startswith("## "):
            headings[2] = stripped[3:].strip()
            headings.pop(3, None)
        elif stripped.startswith("# "):
            headings[1] = stripped[2:].strip()
            headings.pop(2, None)
            headings.pop(3, None)
    parts = [headings[level] for level in sorted(headings.keys())]
    return " > ".join(parts) if parts else ""


def chunk_text(content: str, file_path: Path,
               chunk_size: int = None,
               chunk_overlap: int = None) -> list[str]:
    """Split content into chunks using appropriate splitter for file type."""
    ext = file_path.suffix.lower()
    defaults = FORMAT_CHUNK_DEFAULTS.get(
        ext, {"chunk_size": DEFAULT_CHUNK_SIZE, "chunk_overlap": DEFAULT_CHUNK_OVERLAP}
    )
    chunk_size = chunk_size if chunk_size is not None else defaults["chunk_size"]
    chunk_overlap = chunk_overlap if chunk_overlap is not None else defaults["chunk_overlap"]

    if ext == ".md":
        splitter = MarkdownTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        )
        raw_chunks = splitter.split_text(content)
        enriched = []
        for chunk in raw_chunks:
            # Find where this chunk appears in original content
            search_key = chunk[:80].strip()
            pos = content.find(search_key)
            if pos > 0:
                heading_chain = _get_heading_chain(content, pos)
                if heading_chain and not chunk.strip().startswith("# "):
                    chunk = f"[{heading_chain}]\n\n{chunk}"
            enriched.append(chunk)
        chunks = enriched
    else:
        splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
        )
        chunks = splitter.split_text(content)
    # Filter out very short chunks (less than 50 chars)
    return [c for c in chunks if len(c.strip()) >= 50]


def compute_file_hash(file_path: Path) -> str:
    """Compute MD5 hash of file contents for change detection."""
    return hashlib.md5(file_path.read_bytes()).hexdigest()


def load_hash_cache(cache_path: Path) -> dict:
    """Load the file hash cache for incremental updates."""
    if cache_path.exists():
        return json.loads(cache_path.read_text())
    return {}


def save_hash_cache(cache_path: Path, cache: dict):
    """Save the file hash cache."""
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(json.dumps(cache, indent=2))


def ingest(
    repo_root: Path,
    db_path: Path,
    dry_run: bool = False,
    force: bool = False,
    verbose: bool = False,
    chunk_size: int = DEFAULT_CHUNK_SIZE,
    chunk_overlap: int = DEFAULT_CHUNK_OVERLAP,
):
    """Main ingestion pipeline."""

    print(f"=== Vector Search Ingestion ===")
    print(f"Repo root: {repo_root}")
    print(f"DB path:   {db_path}")
    print(f"Formats:   {', '.join(sorted(e.lstrip('.') for e in SUPPORTED_EXTENSIONS))}")
    print(f"Chunk size: {chunk_size} chars, overlap: {chunk_overlap} chars")
    print()

    # Find all supported files
    all_files = find_files(repo_root)
    print(f"Found {len(all_files)} files")

    if not all_files:
        print("No files to process")
        return

    # Load hash cache for incremental updates
    cache_path = db_path / "file_hashes.json"
    hash_cache = {} if force else load_hash_cache(cache_path)

    # Determine which files need processing
    files_to_process = []
    files_unchanged = 0
    for f in all_files:
        file_hash = compute_file_hash(f)
        rel_path = str(f.relative_to(repo_root))
        if not force and hash_cache.get(rel_path) == file_hash:
            files_unchanged += 1
            if verbose:
                print(f"  SKIP (unchanged): {rel_path}")
        else:
            files_to_process.append((f, file_hash))
            if verbose:
                action = "NEW" if rel_path not in hash_cache else "CHANGED"
                print(f"  {action}: {rel_path}")

    print(f"Files to process: {len(files_to_process)} (unchanged: {files_unchanged})")
    print()

    if dry_run:
        print("=== DRY RUN — no changes made ===")
        total_chunks = 0
        for f, _ in files_to_process:
            try:
                content = extract_text(f)
            except Exception as e:
                print(f"  WARNING: Failed to extract {f.name}: {e}",
                      file=sys.stderr)
                continue
            chunks = chunk_text(content, f, chunk_size, chunk_overlap)
            total_chunks += len(chunks)
            meta = extract_metadata(f, repo_root, content)
            print(f"  {meta['file_path']}: {len(chunks)} chunks, "
                  f"type={meta['file_type']}, area={meta['area']}, "
                  f"title={meta['title']}")
        print(f"\nTotal chunks: {total_chunks}")
        return

    if not files_to_process:
        print("Nothing to update — all files are current")
        return

    # Initialize ChromaDB
    print("Initialising ChromaDB...")
    db_path.mkdir(parents=True, exist_ok=True)
    client = chromadb.PersistentClient(path=str(db_path))
    collection = client.get_or_create_collection(
        name="brain",
        metadata={"hnsw:space": "cosine", "embedding_model": DEFAULT_EMBEDDING_MODEL},
    )

    # Process each file
    total_chunks = 0
    skipped = 0
    start_time = time.time()

    for i, (f, file_hash) in enumerate(files_to_process, 1):
        rel_path = str(f.relative_to(repo_root))

        # Extract text content
        try:
            content = extract_text(f)
        except Exception as e:
            print(f"  WARNING: Failed to extract {rel_path}: {e}",
                  file=sys.stderr)
            skipped += 1
            continue

        metadata = extract_metadata(f, repo_root, content)
        chunks = chunk_text(content, f, chunk_size, chunk_overlap)

        # Prepend document title to chunks for embedding context
        title = metadata["title"]
        if title:
            enriched_chunks = []
            for chunk in chunks:
                # Don't add title if it's already in the first ~50 chars of the chunk
                if title not in chunk[:len(title) + 50]:
                    enriched_chunks.append(f"[{title}]\n\n{chunk}")
                else:
                    enriched_chunks.append(chunk)
            chunks = enriched_chunks

        if not chunks:
            if verbose:
                print(f"  [{i}/{len(files_to_process)}] {rel_path}: no chunks (too short)")
            continue

        # Delete existing chunks for this file (in case of update)
        try:
            existing = collection.get(where={"file_path": rel_path})
            if existing["ids"]:
                collection.delete(ids=existing["ids"])
                if verbose:
                    print(f"  Deleted {len(existing['ids'])} old chunks for {rel_path}")
        except Exception:
            pass  # Collection might be empty

        # Prepare batch data
        ids = []
        documents = []
        metadatas = []

        for j, chunk in enumerate(chunks):
            chunk_id = f"{rel_path}::chunk_{j}"
            ids.append(chunk_id)
            documents.append(chunk)
            metadatas.append({
                **metadata,
                "chunk_index": j,
                "chunk_count": len(chunks),
                "chunk_length": len(chunk),
                "ingested_at": datetime.now().isoformat(),
            })

        # Add to collection
        collection.add(
            ids=ids,
            documents=documents,
            metadatas=metadatas,
        )

        total_chunks += len(chunks)
        # Update hash cache
        hash_cache[rel_path] = file_hash

        if verbose or i % 10 == 0 or i == len(files_to_process):
            print(f"  [{i}/{len(files_to_process)}] {rel_path}: "
                  f"{len(chunks)} chunks")

    elapsed = time.time() - start_time
    save_hash_cache(cache_path, hash_cache)

    # Print summary
    print()
    print(f"=== Ingestion Complete ===")
    print(f"Files processed: {len(files_to_process) - skipped}")
    if skipped:
        print(f"Files skipped (extraction errors): {skipped}")
    print(f"Total chunks added: {total_chunks}")
    print(f"Total chunks in DB: {collection.count()}")
    print(f"Time: {elapsed:.1f}s")
    print(f"DB location: {db_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Ingest files (markdown, PDF, DOCX, XLSX) into vector DB"
    )
    parser.add_argument(
        "repo_root",
        nargs="?",
        default=os.getcwd(),
        help="Root directory to scan (default: cwd)",
    )
    parser.add_argument(
        "--db-path",
        default=None,
        help="Path for ChromaDB storage (default: REPO_ROOT/.vectordb)",
    )
    parser.add_argument("--dry-run", action="store_true", help="Show what would be done")
    parser.add_argument("--force", action="store_true", help="Re-ingest all files")
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose output")
    parser.add_argument("--chunk-size", type=int, default=DEFAULT_CHUNK_SIZE,
                        help=f"Chunk size in chars (default: {DEFAULT_CHUNK_SIZE})")
    parser.add_argument("--chunk-overlap", type=int, default=DEFAULT_CHUNK_OVERLAP,
                        help=f"Chunk overlap in chars (default: {DEFAULT_CHUNK_OVERLAP})")

    args = parser.parse_args()
    repo_root = Path(args.repo_root).resolve()
    db_path = Path(args.db_path) if args.db_path else repo_root / ".vectordb"

    ingest(
        repo_root=repo_root,
        db_path=db_path,
        dry_run=args.dry_run,
        force=args.force,
        verbose=args.verbose,
        chunk_size=args.chunk_size,
        chunk_overlap=args.chunk_overlap,
    )


if __name__ == "__main__":
    main()
